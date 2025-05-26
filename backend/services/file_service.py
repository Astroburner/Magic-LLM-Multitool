# -*- coding: utf-8 -*-
"""
File-Service für die Verarbeitung verschiedener Dateitypen.
"""
import logging
import os
import mimetypes
from typing import Dict, List, Any, Optional
import base64

logger = logging.getLogger(__name__)

# Text-basierte Dateien
TEXT_EXTENSIONS = {'.txt', '.md', '.py', '.js', '.html', '.css', '.json', '.xml', 
                  '.yaml', '.yml', '.log', '.sql', '.sh', '.bat', '.ps1', '.csv'}

def parse_uploaded_files(files: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Verarbeitet hochgeladene Dateien und extrahiert den Inhalt.
    
    Args:
        files: Liste von Datei-Dictionaries mit 'name', 'type', 'data'
        
    Returns:
        Liste von verarbeiteten Dateien mit Inhalt
    """
    processed_files = []
    
    for file_data in files:
        try:
            file_name = file_data.get('name', 'unknown')
            file_type = file_data.get('type', 'unknown')
            file_content_b64 = file_data.get('data', '')
            
            logger.info(f"Verarbeite Datei: {file_name} ({file_type})")
            
            # Base64 dekodieren
            try:
                file_content = base64.b64decode(file_content_b64)
            except Exception as e:
                logger.error(f"Fehler beim Dekodieren von {file_name}: {e}")
                continue
            
            # Dateierweiterung ermitteln
            file_ext = os.path.splitext(file_name)[1].lower()
            
            # Datei verarbeiten basierend auf Typ
            parsed_content = None
            
            if file_ext in TEXT_EXTENSIONS:
                parsed_content = parse_text_file(file_content, file_name)
            elif file_ext == '.pdf':
                parsed_content = parse_pdf_file(file_content, file_name)
            elif file_ext in {'.docx', '.doc'}:
                parsed_content = parse_word_file(file_content, file_name)
            elif file_ext in {'.xlsx', '.xls'}:
                parsed_content = parse_excel_file(file_content, file_name)
            else:
                # Versuche als Text zu lesen
                parsed_content = parse_text_file(file_content, file_name)
            
            if parsed_content:
                processed_files.append({
                    'name': file_name,
                    'type': file_type,
                    'extension': file_ext,
                    'content': parsed_content,
                    'size': len(file_content)
                })
                
        except Exception as e:
            logger.error(f"Fehler beim Verarbeiten der Datei {file_name}: {e}")
            continue
    
    logger.info(f"Erfolgreich {len(processed_files)} Dateien verarbeitet")
    return processed_files

def parse_text_file(content: bytes, filename: str) -> Optional[str]:
    """Verarbeitet Text-Dateien."""
    try:
        # Verschiedene Encodings versuchen
        for encoding in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue
        
        logger.warning(f"Konnte Encoding für {filename} nicht ermitteln")
        return content.decode('utf-8', errors='replace')
        
    except Exception as e:
        logger.error(f"Fehler beim Lesen der Text-Datei {filename}: {e}")
        return None

def parse_pdf_file(content: bytes, filename: str) -> Optional[str]:
    """Verarbeitet PDF-Dateien."""
    try:
        import PyPDF2
        import io
        
        # PDF aus Bytes lesen
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(content))
        
        text_content = []
        for page_num, page in enumerate(pdf_reader.pages):
            try:
                page_text = page.extract_text()
                if page_text.strip():
                    text_content.append(f"--- Page {page_num + 1} ---\n{page_text}")
            except Exception as e:
                logger.warning(f"Fehler beim Lesen von Seite {page_num + 1} in {filename}: {e}")
                continue
        
        if text_content:
            result = "\n\n".join(text_content)
            logger.info(f"PDF erfolgreich verarbeitet: {filename} ({len(pdf_reader.pages)} Seiten)")
            return result
        else:
            return f"[PDF-Datei: {filename} - Kein Text extrahierbar (möglicherweise nur Bilder)]"
            
    except ImportError:
        logger.error("PyPDF2 nicht installiert - PDF-Parsing nicht verfügbar")
        return f"[PDF-Datei: {filename} - PyPDF2 Library fehlt]"
    except Exception as e:
        logger.error(f"Fehler beim Lesen der PDF-Datei {filename}: {e}")
        return f"[PDF-Datei: {filename} - Fehler beim Parsen: {str(e)}]"

def parse_word_file(content: bytes, filename: str) -> Optional[str]:
    """Verarbeitet Word-Dateien."""
    try:
        from docx import Document
        import io
        
        # Word-Dokument aus Bytes lesen
        doc = Document(io.BytesIO(content))
        
        text_content = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text_content.append(paragraph.text)
        
        if text_content:
            result = "\n\n".join(text_content)
            logger.info(f"Word-Dokument erfolgreich verarbeitet: {filename}")
            return result
        else:
            return f"[Word-Datei: {filename} - Kein Text gefunden]"
            
    except ImportError:
        logger.error("python-docx nicht installiert - Word-Parsing nicht verfügbar")
        return f"[Word-Datei: {filename} - python-docx Library fehlt]"
    except Exception as e:
        logger.error(f"Fehler beim Lesen der Word-Datei {filename}: {e}")
        return f"[Word-Datei: {filename} - Fehler beim Parsen: {str(e)}]"

def parse_excel_file(content: bytes, filename: str) -> Optional[str]:
    """Verarbeitet Excel-Dateien."""
    try:
        import openpyxl
        import io
        
        # Excel-Datei aus Bytes lesen
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        
        text_content = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_content.append(f"--- Sheet: {sheet_name} ---")
            
            # Zelldaten lesen (erste 50 Zeilen)
            rows_data = []
            for row_num, row in enumerate(sheet.iter_rows(max_row=50, values_only=True), 1):
                if any(cell is not None for cell in row):
                    row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    rows_data.append(f"Row {row_num}: {row_str}")
            
            text_content.append("\n".join(rows_data))
            
            if sheet.max_row > 50:
                text_content.append(f"\n... ({sheet.max_row - 50} more rows)")
        
        result = "\n\n".join(text_content)
        logger.info(f"Excel-Datei erfolgreich verarbeitet: {filename}")
        return result
        
    except ImportError:
        logger.error("openpyxl nicht installiert - Excel-Parsing nicht verfügbar")
        return f"[Excel-Datei: {filename} - openpyxl Library fehlt]"
    except Exception as e:
        logger.error(f"Fehler beim Lesen der Excel-Datei {filename}: {e}")
        return f"[Excel-Datei: {filename} - Fehler beim Parsen: {str(e)}]"

def format_files_for_llm(files: List[Dict[str, Any]]) -> str:
    """
    Formatiert verarbeitete Dateien für das LLM.
    
    Args:
        files: Liste verarbeiteter Dateien
        
    Returns:
        Formatierter String für das LLM
    """
    if not files:
        return ""
    
    formatted_content = "\n=== UPLOADED FILES ===\n"
    
    for i, file_data in enumerate(files, 1):
        name = file_data.get('name', 'Unknown')
        content = file_data.get('content', '')
        size = file_data.get('size', 0)
        ext = file_data.get('extension', '')
        
        formatted_content += f"\n--- File {i}: {name} ({ext.upper()}, {size} bytes) ---\n"
        formatted_content += content[:5000]  # Limit auf 5000 Zeichen pro Datei
        
        if len(content) > 5000:
            formatted_content += "\n... [Content truncated] ..."
        
        formatted_content += "\n" + "-" * 50 + "\n"
    
    formatted_content += "\n=== END FILES ===\n"
    return formatted_content